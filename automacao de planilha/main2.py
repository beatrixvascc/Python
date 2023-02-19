import openpyxl

planilha = openpyxl.load_workbook("Planilha de clientes.xlsx")   #parte 1 - carregando planilha 
dashboard = planilha["Clientes"]    #parte 2 - escolher a pagina
for linhas in dashboard.iter_rows(min_row=2,max_row=5):   #parte 3 - percorre as linhas
    print(f"{linhas[0].value},{linhas[1].value},{linhas[2].value}")
planilha.save("Planilha de clientes 2.xlsx")




#se quiser alterar:
'''
for linhas in dashboard.iter_rows(min_row=2,max_row=5):
    for celula in linhas:
        if celula.value=="analise":
            celula.value="venda feita"
'''            
